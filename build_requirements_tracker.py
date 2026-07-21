"""
Generates Requirements_Tracker.xlsx — a simple, shareable reference for what
data this project generates, story by story, plus a visual architecture
diagram, a plain-language glossary, and a known-issues tracker.

Design goal: readable by anyone in 30 seconds. No complex numbering, no
formula-heavy dashboards, no jargon-heavy walls of text.

Usage:
    py build_requirements_tracker.py

Output:
    Requirements_Tracker.xlsx (repo root)
"""
from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

TODAY = date.today().isoformat()
OUT_PATH = "Requirements_Tracker.xlsx"

# ---------------------------------------------------------------------------
# Palette / styles
# ---------------------------------------------------------------------------
NAVY = "1F3864"
BLUE = "2F5597"
LIGHT_BLUE = "D9E2F3"
TEAL = "1F7A6C"
LIGHT_TEAL = "D7F0EC"
PLUM = "6B3FA0"
LIGHT_PLUM = "E7DBF3"
ORANGE = "C55A11"
LIGHT_ORANGE = "FCE4D6"
GRAY = "F2F2F2"
WHITE = "FFFFFF"
GREEN = "C6EFCE"
GREEN_TXT = "006100"
YELLOW = "FFEB9C"
YELLOW_TXT = "9C6500"
RED = "FFC7CE"
RED_TXT = "9C0006"
GRAY_TXT = "606060"

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color=GRAY_TXT)
SECTION_FONT = Font(name="Calibri", size=13, bold=True, color=WHITE)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_BODY = Font(name="Calibri", size=10, bold=True)

HEADER_FILL = PatternFill("solid", fgColor=BLUE)
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
STRIPE_FILL = PatternFill("solid", fgColor=GRAY)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
CENTER = Alignment(vertical="center", horizontal="center")


def style_header_row(ws, row, ncols, height=28):
    ws.row_dimensions[row].height = height
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = BORDER


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_status_colors(ws, col_letter, grow_to=1000):
    rng = f"{col_letter}2:{col_letter}{grow_to}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Implemented"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_TXT)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fixed"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_TXT)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"In Progress"'], fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_TXT)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Open"'], fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_TXT)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Not Started"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_TXT)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Blocked"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_TXT)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Context Only"'], fill=PatternFill("solid", fgColor=GRAY), font=Font(color=GRAY_TXT, italic=True)))


# ---------------------------------------------------------------------------
# Story Overview + Variables (unchanged content from prior version)
# ---------------------------------------------------------------------------
CORE_DOC = "2026.07.10 Synthetic Wafer Map Generator.docx"
STORY_DOC = "2026.07.14 Yield Stories for Synthetic Data.pdf"

STORY_OVERVIEW = [
    ("Core Wafer Map Generation", CORE_DOC,
     "Generates the base synthetic wafer for a lot: die grid, spatial pass/fail pattern, yield, "
     "CP/FT bin data, and test item values. Every other story builds on top of this.",
     "Implemented", "geometry.py, signatures.py, yield_model.py, binning.py, test_items.py, fab.py, generator.py"),
    ("ECID Matching / Traceability", STORY_DOC,
     "Generates matched Circuit Probe (CP) and Final Test (FT) datasets linked by a unique chip ID (ECID), "
     "including intentional traceability breaks (blank IDs, wrong picks, sweeper lots) so CP-to-FT-to-field "
     "matching — and its failure modes — can be demoed.",
     "Implemented", "ecid.py, assembly.py, gdbn.py, multidie.py, final_test.py, story1_presets.py"),
    ("Cluster Tool Yield Problem", STORY_DOC,
     "Generates lots where one chamber of a multi-chamber fab tool is bad, lowering yield only for the "
     "wafers that chamber processed — used to demo detecting a tool-level (not lot-level) yield problem.",
     "Not Started", "—"),
    ("Fab Equipment Commonality", STORY_DOC,
     "Generates a time series of fab lots where one piece of lot-level equipment intermittently causes low "
     "yield — invisible to a fabless customer (looks like random bad lots) or traceable via equipment/FDC "
     "history for an IDM/Foundry.",
     "Not Started", "—"),
    ("Sort Fixture Wear-out", STORY_DOC,
     "Generates a wafer where yield degrades over the course of testing due to probe-tip wear, either "
     "permanently or with periodic recovery from automatic probe cleaning.",
     "Not Started", "—"),
    ("Non-Normal Yield Variation", STORY_DOC,
     "Generates a lot where one or more wafers behave very differently from the rest (different spatial "
     "pattern + worse defect density) — used to demo outlier detection.",
     "Not Started", "—"),
]
OVERVIEW_HEADERS = ["Story", "Source Doc", "What Data It Generates", "Status", "Key Files"]

VARIABLES = [
    ("Core Wafer Map Generation", "Wafer Diameter", "—", "Overall wafer size, which sets the die grid boundary", "300 mm", "Implemented", "geometry.py"),
    ("Core Wafer Map Generation", "Edge Exclusion (EE)", "Edge Exclusion", "Ring near the wafer edge with no usable die", "1–10 mm", "Implemented", "geometry.py"),
    ("Core Wafer Map Generation", "Die Size", "—", "Width x height of each chip; drives how many die fit per wafer (GDPW)", "1x1–25x35 mm", "Implemented", "geometry.py"),
    ("Core Wafer Map Generation", "Scribe Street", "—", "Gap left between neighboring die", "0.05–0.2 mm", "Implemented", "geometry.py"),
    ("Core Wafer Map Generation", "Flat / Notch", "—", "Wafer edge marking style, auto-chosen from diameter", "flat (150mm) / notch (200/300mm)", "Implemented", "geometry.py"),
    ("Core Wafer Map Generation", "Wafer Orientation", "—", "Which side the flat/notch faces", "down/up/left/right", "Implemented", "geometry.py"),
    ("Core Wafer Map Generation", "Wafer Quantity", "—", "Number of wafers generated per lot", "1–25", "Implemented", "llm_agent.py"),
    ("Core Wafer Map Generation", "Yield (Y)", "—", "% of die that pass test on the wafer", "direct % or derived from D0", "Implemented", "yield_model.py"),
    ("Core Wafer Map Generation", "D0", "Defect Density", "Defects per cm² used to derive yield: Y = e^(-A x D0)", "0.08 defects/cm²", "Implemented", "yield_model.py, geometry.py"),
    ("Core Wafer Map Generation", "CP1 / CP2 / CP3", "Circuit Probe (test insertions)", "Which wafer-probe test stages run, each keeping a % of the prior stage's passers", "CP1 only, or CP1+CP2, or CP1+CP2+CP3", "Implemented", "yield_model.py"),
    ("Core Wafer Map Generation", "Hardbin / Softbin Count", "—", "How many bin categories test results are grouped into", "hardbin 16/64/256, softbin x4/x16/x64", "Implemented", "binning.py"),
    ("Core Wafer Map Generation", "Test Item Count", "—", "Number of parametric/pass-fail test measurements generated per die", "100 – 1,000,000", "Implemented", "test_items.py"),
    ("Core Wafer Map Generation", "Spatial Signature", "—", "Which named fail pattern is drawn on the wafer (edge ring, donut, scratch, cluster, etc. — 37 total)", "—", "Implemented", "signatures.py"),
    ("Core Wafer Map Generation", "Repeater / Reticle Pattern", "—", "A fail that repeats at the same die position in every stepper field", "hard or soft, 10–100% fail rate", "Implemented", "signatures.py"),
    ("Core Wafer Map Generation", "Striping", "—", "Fail pattern along one edge of the stepper field (lens tilt)", "top/bottom/left/right", "Implemented", "signatures.py"),
    ("Core Wafer Map Generation", "Lot ID", "Fab + Year + Week + Sequence (FYYWWSSSS)", "Standard lot number format for generated lots", "—", "Implemented", "fab.py"),
    ("Core Wafer Map Generation", "Multi-Site Count", "—", "Number of parallel test sites, auto-derived from die count per wafer (GDPW)", "1/2/4/8/16", "Implemented", "fab.py"),
    ("Core Wafer Map Generation", "S2S Loss", "Site-to-Site (yield loss)", "One probe site performs worse than the others", "healthy >95%, problem site 40–80%", "Implemented", "yield_model.py"),

    ("ECID Matching / Traceability", "ECID", "Electronic Chip ID", "Unique ID burned into each die at CP, used to match it to its FT (final test) result", "—", "Implemented", "ecid.py"),
    ("ECID Matching / Traceability", "ECID Encoding Mode", "—", "How the ECID value is represented: plain (lot+wafer+x+y), ROT13 'encrypted', or split into 4 separate test items", "plain", "Implemented", "ecid.py"),
    ("ECID Matching / Traceability", "Traceability Case", "—", "Which CP-to-FT packaging scenario is generated: single die (Case A), multi-die (Case B), or IDM/Foundry with factory data (Case C/D)", "Case A", "Implemented (A, B); Case C/D deferred by spec ('TK')", "multidie.py, assembly.py"),
    ("ECID Matching / Traceability", "Multi-Die Mode", "—", "For multi-die products: whether every component chip is traceable (Full) or one is not (Partial)", "Full", "Implemented", "multidie.py"),
    ("ECID Matching / Traceability", "Assembly Scenario", "—", "How CP die map to FT lots: one-to-one, or a 'sweeper' lot combining passers from multiple CP lots", "one-to-one", "Implemented", "assembly.py"),
    ("ECID Matching / Traceability", "Assembly Error Type", "—", "Simulated assembly mistake: wrong bin picked, or wrong X/Y coordinate picked", "none", "Implemented", "assembly.py"),
    ("ECID Matching / Traceability", "Blank ECID Rate", "—", "% of FT units with no ECID (simulates power-shorted / cracked die that never wrote an ID)", "<2%", "Implemented", "assembly.py"),
    ("ECID Matching / Traceability", "GDBN Mode", "Good Die Bad Neighborhood", "Models low yield at FT caused by CP test clusters that were missed by loose CP limits", "off", "Implemented", "gdbn.py"),
    ("ECID Matching / Traceability", "GDBN Neighbor Fail Rate", "—", "Chance that a good die next to a real CP fail cluster still fails at FT", "50%", "Implemented", "gdbn.py"),

    ("Cluster Tool Yield Problem", "Chamber Count", "—", "Number of parallel process chambers in the fab tool", "3", "Not Started", "—"),
    ("Cluster Tool Yield Problem", "Bad Chamber Count", "—", "How many of the chambers are bad at once", "1 (fixed, per spec)", "Not Started", "—"),
    ("Cluster Tool Yield Problem", "Bad Chamber Selection", "—", "Which chamber is bad", "random", "Not Started", "—"),
    ("Cluster Tool Yield Problem", "Yield Impact", "—", "How much worse yield is on wafers processed by the bad chamber", "-10% (90% of normal)", "Not Started", "—"),
    ("Cluster Tool Yield Problem", "FT-Only Variant", "—", "Whether the problem is invisible at CP and only shows up at FT (traced via ECID)", "off (nice-to-have)", "Not Started", "—"),

    ("Fab Equipment Commonality", "Lot Time Series Count", "—", "How many fab lots are generated in the demo time series", "20 lots x 25 wafers", "Not Started", "—"),
    ("Fab Equipment Commonality", "D0 Baseline + Spread", "Defect Density", "Nominal defect density and lot-to-lot variation for 'good' lots", "0.08 defects/cm² ± 10%", "Not Started", "—"),
    ("Fab Equipment Commonality", "Die Size Range", "—", "Range of die sizes used, to vary die-per-wafer (GDPW) across the demo", "58–994 GDPW (default ~101 GDPW)", "Not Started", "—"),
    ("Fab Equipment Commonality", "Bad Lot Percentage", "—", "Fraction of lots affected by the bad equipment", "20%", "Not Started", "—"),
    ("Fab Equipment Commonality", "Bad Lot D0 Multiplier", "Defect Density", "How much worse defect density is on bad lots vs. baseline", "3x baseline", "Not Started", "—"),
    ("Fab Equipment Commonality", "Visibility Mode", "—", "Fabless (equipment problem invisible, looks random) vs IDM/Foundry (traceable to Lot Equipment History / FDC data)", "Fabless", "Not Started (Fabless is must-have; IDM/Foundry is nice-to-have)", "—"),

    ("Sort Fixture Wear-out", "Degradation Mode", "—", "Whether yield degradation recovers periodically (Complex) or not at all (Simple)", "Simple", "Not Started", "—"),
    ("Sort Fixture Wear-out", "Starting Yield", "—", "Nominal wafer yield and spread before degradation begins", "N% ± M%", "Not Started", "—"),
    ("Sort Fixture Wear-out", "Degradation Multiplier (P)", "—", "Final yield multiplier the wafer trends toward as probe wear worsens", "0.6", "Not Started", "—"),
    ("Sort Fixture Wear-out", "Touchdowns to Bottom (Q)", "—", "How many die get tested before yield bottoms out", "500 die", "Not Started", "—"),
    ("Sort Fixture Wear-out", "Reset Interval (R)", "—", "Touchdowns between automatic probe-cleaning resets (Complex mode only)", "50 die", "Not Started", "—"),

    ("Non-Normal Yield Variation", "Outlier Wafer Count", "—", "How many wafers in the lot behave like outliers", "1 to N", "Not Started", "—"),
    ("Non-Normal Yield Variation", "Outlier Spatial Pattern", "—", "The distinct fail pattern given to outlier wafers, different from the rest of the lot", "—", "Not Started", "—"),
    ("Non-Normal Yield Variation", "Yield Impact Multiplier (M)", "—", "How much worse defect density is on outlier wafers: D0(outlier) = D0(rest) x M", "—", "Not Started", "—"),
    ("Non-Normal Yield Variation", "Wafer Selection Mode", "—", "Which wafer(s) become outliers: 1st only, last only, random 1-3, or a sequential run (e.g. wafers 1-6)", "—", "Not Started", "—"),
    ("Non-Normal Yield Variation", "Parametric Variation", "—", "Non-normal variation applied to parametric (not just pass/fail) data", "—", "Deferred by spec ('TK' — no detail given yet)", "—"),
]
VARIABLE_HEADERS = ["Story", "Variable", "Full Form", "What It Generates / Controls", "Default / Typical Value", "Status", "Implemented In"]

STATUS_LIST = "Not Started,In Progress,Implemented,Context Only"


def build_story_overview(wb):
    ws = wb.active
    ws.title = "Story Overview"
    ncols = len(OVERVIEW_HEADERS)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"] = "Wafer Map Generator — What This Project Generates, By Story"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    ws["A2"] = (
        f"Goal: synthetic wafer-map / test data for PDF Solutions demos — no real customer data needed. "
        f"Last generated: {TODAY}. Full variable list: 'Variables' tab. How it's built: 'Technical Architecture' tab. "
        f"Known issues / needed changes: 'Bugs & Needs Changing' tab."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 16

    header_row = 4
    for i, h in enumerate(OVERVIEW_HEADERS, start=1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, ncols, height=24)

    for offset, (story, doc, desc, status, files) in enumerate(STORY_OVERVIEW):
        r = header_row + 1 + offset
        values = [story, doc, desc, status, files]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BOLD_BODY if c == 1 else BODY_FONT
            cell.border = BORDER
            cell.alignment = WRAP
            if offset % 2 == 1:
                cell.fill = STRIPE_FILL

    last_row = header_row + len(STORY_OVERVIEW)
    add_status_colors(ws, "D")

    ws.freeze_panes = f"A{header_row + 1}"
    autosize(ws, {"A": 26, "B": 34, "C": 62, "D": 13, "E": 44})
    for r in range(header_row + 1, last_row + 1):
        ws.row_dimensions[r].height = 46
    ws.sheet_view.showGridLines = False
    return ws


def build_variables(wb):
    ws = wb.create_sheet("Variables")
    ncols = len(VARIABLE_HEADERS)

    for i, h in enumerate(VARIABLE_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, ncols, height=26)

    for r, row in enumerate(VARIABLES, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BOLD_BODY if c == 2 else BODY_FONT
            cell.border = BORDER
            cell.alignment = WRAP

    last_row = len(VARIABLES) + 1
    table_ref = f"A1:{get_column_letter(ncols)}{last_row}"
    table = Table(displayName="Variables", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    dv = DataValidation(type="list", formula1=f'"{STATUS_LIST}"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add("F2:F1000")

    add_status_colors(ws, "F")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = table_ref
    autosize(ws, {"A": 26, "B": 24, "C": 34, "D": 50, "E": 26, "F": 15, "G": 30})
    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# Technical Architecture — drawn as a block diagram (boxes + arrows)
# ---------------------------------------------------------------------------
DIAGRAM_COLS = ("A", "B", "C", "D", "E", "F")
COL1, COL2 = 1, 6


def draw_box(ws, top_row, title, body, dark, light, title_size=11):
    """Draws one diagram box: a dark title bar + a light body area, bordered
    as a single unit. Returns the row number just below the box."""
    ws.merge_cells(start_row=top_row, start_column=COL1, end_row=top_row, end_column=COL2)
    tcell = ws.cell(row=top_row, column=COL1, value=title)
    tcell.font = Font(name="Calibri", size=title_size, bold=True, color=WHITE)
    tcell.fill = PatternFill("solid", fgColor=dark)
    tcell.alignment = WRAP_CENTER
    ws.row_dimensions[top_row].height = 22

    body_top = top_row + 1
    n_lines = body.count("\n") + 1 if body else 1
    body_bottom = body_top + n_lines - 1
    ws.merge_cells(start_row=body_top, start_column=COL1, end_row=body_bottom, end_column=COL2)
    bcell = ws.cell(row=body_top, column=COL1, value=body)
    bcell.font = Font(name="Calibri", size=10.5, color=NAVY)
    bcell.fill = PatternFill("solid", fgColor=light)
    bcell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
    for rr in range(body_top, body_bottom + 1):
        ws.row_dimensions[rr].height = 18

    thick = Side(style="medium", color=dark)
    box_border = Border(left=thick, right=thick, top=thick, bottom=thick)
    for rr in range(top_row, body_bottom + 1):
        for cc in range(COL1, COL2 + 1):
            ws.cell(row=rr, column=cc).border = box_border

    return body_bottom + 1


def draw_arrow(ws, row, label=""):
    ws.merge_cells(start_row=row, start_column=COL1, end_row=row, end_column=COL2)
    cell = ws.cell(row=row, column=COL1, value=("↓  " + label if label else "↓"))
    cell.font = Font(name="Calibri", size=14, bold=True, color=NAVY)
    cell.alignment = CENTER
    ws.row_dimensions[row].height = 22
    return row + 1


def draw_gap(ws, row, height=8):
    ws.row_dimensions[row].height = height
    return row + 1


def build_architecture(wb):
    ws = wb.create_sheet("Technical Architecture")
    r = 1
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"] = "Technical Architecture — How It's Built"
    ws[f"A{r}"].font = TITLE_FONT
    ws[f"A{r}"].fill = TITLE_FILL
    ws[f"A{r}"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 32
    r += 1

    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"] = "In plain terms: you describe a scenario, one file turns it into wafer data, and files come out. Diagram below, top to bottom."
    ws[f"A{r}"].font = SUBTITLE_FONT
    r += 2

    # --- Diagram 1: request -> wafer data ---
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"] = "1) How Your Request Becomes Wafer Data"
    ws[f"A{r}"].font = SECTION_FONT
    ws[f"A{r}"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"A{r}"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 22
    r += 1
    r = draw_gap(ws, r)

    r = draw_box(
        ws, r,
        "STEP 1 — You Describe a Scenario",
        "Fill out a form, pick a ready-made Story, or type a request in plain English.\n"
        "All three do the exact same thing underneath.",
        BLUE, LIGHT_BLUE,
    )
    r = draw_arrow(ws, r)

    r = draw_box(
        ws, r,
        "STEP 2 — generator.py Builds the Wafer(s)",
        "One file runs every scenario, always in this order:\n"
        "  1. Shape & grid — geometry.py\n"
        "  2. Fail pattern — signatures.py\n"
        "  3. Yield & retest — yield_model.py\n"
        "  4. Bins — binning.py\n"
        "  5. Traceability, ECID story only — ecid.py, assembly.py, gdbn.py, multidie.py",
        TEAL, LIGHT_TEAL,
    )
    r = draw_arrow(ws, r)

    r = draw_box(
        ws, r,
        "STEP 3 — Files Come Out",
        "CSV data, wafer map pictures (PNG/SVG/JPEG/TIFF), and STDF files —\n"
        "ready to load into a demo tool like Exensio.",
        ORANGE, LIGHT_ORANGE,
    )
    r += 1
    r = draw_gap(ws, r, height=14)

    # --- Diagram 2: where the code lives ---
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"] = "2) Where the Code Lives"
    ws[f"A{r}"].font = SECTION_FONT
    ws[f"A{r}"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"A{r}"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 22
    r += 1
    r = draw_gap(ws, r)

    r = draw_box(
        ws, r,
        "SCREEN — What You See and Type",
        "app.py — the web page (form, Stories tab, chat box)\n"
        "llm_agent.py — reads what you typed and figures out what you want",
        BLUE, LIGHT_BLUE,
    )
    r = draw_arrow(ws, r)

    r = draw_box(
        ws, r,
        "BRAIN — Turns Requests Into Data",
        "generator.py — the pipeline described above",
        NAVY, LIGHT_BLUE,
    )
    r = draw_arrow(ws, r)

    r = draw_box(
        ws, r,
        "BUILDING BLOCKS — Core Wafer Rules",
        "geometry.py (shape/grid) · signatures.py (fail patterns) · yield_model.py (yield & retest)\n"
        "binning.py (bins) · test_items.py (test data) · fab.py (lot IDs, timing, multi-site)",
        TEAL, LIGHT_TEAL,
    )
    r = draw_arrow(ws, r, "plus, for the ECID story:")

    r = draw_box(
        ws, r,
        "TRACEABILITY ADD-ON — ECID Story",
        "ecid.py (chip ID) · assembly.py (CP-to-FT matching) · gdbn.py (bad-neighborhood yield loss)\n"
        "multidie.py (multi-chip products) · final_test.py (final results) · story1_presets.py (defaults)",
        PLUM, LIGHT_PLUM,
    )
    r = draw_arrow(ws, r)

    r = draw_box(
        ws, r,
        "EXPORT — Turns Data Into Files",
        "renderer.py (draws the wafer map pictures) · stdf_writer.py (writes STDF files)",
        ORANGE, LIGHT_ORANGE,
    )
    r += 1
    r = draw_gap(ws, r, height=14)

    r = draw_box(
        ws, r,
        "TESTS — Makes Sure Nothing Breaks",
        "The tests/ folder automatically checks the tricky ECID Story cases every time the code changes.",
        GRAY_TXT, GRAY,
        title_size=10,
    )
    r += 1
    r = draw_gap(ws, r, height=14)

    # --- Section 3: tools we use ---
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"] = "3) Tools We Use"
    ws[f"A{r}"].font = SECTION_FONT
    ws[f"A{r}"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"A{r}"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 22
    r += 1
    r = draw_gap(ws, r)

    r = draw_box(
        ws, r,
        "Tech Stack, In Plain English",
        "Python — the programming language everything is written in\n"
        "Streamlit — makes the web page you interact with\n"
        "pandas / numpy — crunch the numbers behind the scenes\n"
        "matplotlib — draws the wafer map pictures\n"
        "Azure OpenAI (GPT-4.1) — reads your typed requests in the chat tab\n"
        "pytest — automatically checks the code still works\n"
        "Streamlit Community Cloud — hosts the live web page for sharing",
        BLUE, LIGHT_BLUE,
    )
    r += 1
    r = draw_gap(ws, r, height=14)

    # --- Section 4: adding a new story ---
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"] = "4) Adding a New Story — Checklist"
    ws[f"A{r}"].font = SECTION_FONT
    ws[f"A{r}"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"A{r}"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 22
    r += 1
    r = draw_gap(ws, r)

    r = draw_box(
        ws, r,
        "When a new Story spec arrives, do these in order:",
        "1. Add its variables to the Variables tab (even before building anything)\n"
        "2. Write one new file for its logic, e.g. gdbn.py\n"
        "3. Add a presets file with its default settings\n"
        "4. Plug the new file into generator.py\n"
        "5. Add controls for it on the Stories tab in app.py\n"
        "6. Write tests for the tricky cases the spec calls out\n"
        "7. Flip its Status to Implemented on the Story Overview / Variables tabs",
        TEAL, LIGHT_TEAL,
    )

    ws.sheet_view.showGridLines = False
    autosize(ws, {"A": 16, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
    ws.freeze_panes = "A3"
    return ws


# ---------------------------------------------------------------------------
# Glossary — plain-language, alphabetical
# ---------------------------------------------------------------------------
GLOSSARY = [
    ("ALPS", "A data source that tracks which fab machine touched which wafer/lot. Needed for advanced equipment-tracing — not built yet."),
    ("Cluster tool", "One fab machine with 2 or more processing chambers that work on wafers at the same time."),
    ("CP", "Circuit Probe — testing the chips while they're still on the wafer, before it's cut apart."),
    ("D0", "Defect Density — defects per square centimeter. Higher D0 means lower yield."),
    ("ECID", "Electronic Chip ID — a unique number written onto each chip so it can be matched from wafer test to final test."),
    ("FDC", "Fault Detection and Classification — data fabs use to watch equipment health."),
    ("FOUP", "The container that holds and moves wafers around the fab."),
    ("FT", "Final Test — testing the finished, packaged chip (happens after CP)."),
    ("GDBN", "Good Die Bad Neighborhood — a chip that passed wafer test, but is sitting right next to bad chips and later fails final test."),
    ("GDPW", "Gross Die Per Wafer — how many chips fit on one wafer."),
    ("Hardbin / Softbin", "Two ways of grouping test results. Hardbin is the simple, coarse category; softbin is more detailed."),
    ("Lot ID", "A lot's tracking number, formatted as Fab letter + Year + Week + Sequence number (e.g. F26289001)."),
    ("Multi-die product", "A finished chip package built from more than one die, e.g. a logic chip and a memory chip packaged together."),
    ("Reticle / Repeater", "A defect that shows up in the same spot on every stepper exposure, so it repeats in a grid pattern across the wafer."),
    ("S2S", "Site-to-Site — when one test site, in a multi-site test, gives worse results than the others."),
    ("STDF", "Standard Test Data Format — the standard file type test data is saved in, so tools like Exensio can read it."),
    ("Sweeper lot", "A final-test lot built by combining passing chips from several different wafer-test lots."),
]


def build_glossary(wb):
    ws = wb.create_sheet("Glossary")
    ws.merge_cells("A1:B1")
    ws["A1"] = "Glossary — Plain-Language Terms"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.cell(row=3, column=1, value="Term")
    ws.cell(row=3, column=2, value="What It Means")
    style_header_row(ws, 3, 2, height=20)

    for i, (term, defn) in enumerate(GLOSSARY, start=4):
        tc = ws.cell(row=i, column=1, value=term)
        dc = ws.cell(row=i, column=2, value=defn)
        tc.font = BOLD_BODY
        dc.font = BODY_FONT
        tc.alignment = WRAP
        dc.alignment = WRAP
        tc.border = BORDER
        dc.border = BORDER
        if (i - 4) % 2 == 1:
            tc.fill = STRIPE_FILL
            dc.fill = STRIPE_FILL

    ws.freeze_panes = "A4"
    autosize(ws, {"A": 22, "B": 95})
    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# Bugs & Needs Changing — known issues in what's already built
# ---------------------------------------------------------------------------
BUGS_HEADERS = ["Area / File", "What It Is Now", "What Needs Changing", "Type", "Priority", "Status"]
TYPE_LIST = "Bug,Missing Feature,Enhancement,Needs Verification,Blocked"
PRIORITY_LIST = "High,Medium,Low"
BUG_STATUS_LIST = "Open,In Progress,Fixed,Blocked,Won't Fix"

BUGS = [
    ("test_items.py", "Parametric test values are random within a range (uniform distribution).",
     "Should look more like real test data — use a bell-curve (Gaussian) distribution instead.",
     "Enhancement", "Medium", "Open"),
    ("signatures.py", "37 named fail patterns exist and are selectable.",
     "Not yet checked against Steve's official list of patterns — some names or shapes may not match.",
     "Needs Verification", "Medium", "Open"),
    ("fab.py", "Code can already create split-lot numbers (e.g. LOT.01, LOT.02).",
     "Not wired up to any screen yet — there's no way to actually request one.",
     "Missing Feature", "Low", "Open"),
    ("Repair (virgin vs repaired die)", "Not built.",
     "Need to decide how a 'repaired good' die should differ from a 'virgin good' die before building it. Spec itself says this needs more thought.",
     "Missing Feature", "Low", "Open"),
    ("ECID Story — IDM/Foundry case", "Not built.",
     "Needs real factory/equipment (ALPS) data before this can be built. Blocked until that data exists.",
     "Blocked", "Low", "Blocked"),
    ("multidie.py", "Each chip in a multi-chip product is generated as a simple pass/fail chance, not a full wafer map.",
     "Matches the spec's own 'simple case' guidance for now — revisit only if a demo needs full per-component wafer maps.",
     "Enhancement", "Low", "Open"),
]


def build_bugs(wb):
    ws = wb.create_sheet("Bugs & Needs Changing")
    ncols = len(BUGS_HEADERS)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"] = "Bugs & Needs Changing"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    ws["A2"] = (
        "Scope: known issues and needed changes in what's already built. "
        "Stories that haven't been started yet are tracked on the 'Variables' tab instead."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 16

    header_row = 4
    for i, h in enumerate(BUGS_HEADERS, start=1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, ncols, height=24)

    for r, row in enumerate(BUGS, start=header_row + 1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BOLD_BODY if c == 1 else BODY_FONT
            cell.border = BORDER
            cell.alignment = WRAP

    last_row = header_row + len(BUGS)
    table_ref = f"A{header_row}:{get_column_letter(ncols)}{last_row}"
    table = Table(displayName="BugsList", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    grow_to = 500
    dv_type = DataValidation(type="list", formula1=f'"{TYPE_LIST}"', allow_blank=True, showDropDown=False)
    dv_priority = DataValidation(type="list", formula1=f'"{PRIORITY_LIST}"', allow_blank=True, showDropDown=False)
    dv_status = DataValidation(type="list", formula1=f'"{BUG_STATUS_LIST}"', allow_blank=True, showDropDown=False)
    for dv in (dv_type, dv_priority, dv_status):
        ws.add_data_validation(dv)
    dv_type.add(f"D{header_row + 1}:D{grow_to}")
    dv_priority.add(f"E{header_row + 1}:E{grow_to}")
    dv_status.add(f"F{header_row + 1}:F{grow_to}")

    add_status_colors(ws, "F")

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = table_ref
    autosize(ws, {"A": 26, "B": 44, "C": 48, "D": 18, "E": 10, "F": 12})
    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# Change Log
# ---------------------------------------------------------------------------
CHANGELOG = [
    (TODAY, "All", "Simplified workbook: replaced numbered Dashboard/Requirements Detail with plain 'Story Overview' + 'Variables' tabs.", "—", "Auto-generated"),
    (TODAY, "Technical Architecture", "Rebuilt as a visual block diagram (boxes + arrows) with plain-language wording instead of a text list.", "—", ""),
    (TODAY, "Glossary", "Simplified definitions to plain language and sorted alphabetically.", "—", ""),
    (TODAY, "Bugs & Needs Changing", "New tab added to track known issues and needed changes in what's already built.", "—", ""),
    (TODAY, "Stories 2-5", "Added Cluster Tool, Fab Equipment Commonality, Sort Fixture Wear-out, and Non-Normal Yield Variation as Not Started.", "2026.07.14 Yield Stories for Synthetic Data.pdf", ""),
]
CHANGELOG_HEADERS = ["Date", "Area", "Change Description", "Source Doc", "Notes"]


def build_changelog(wb):
    ws = wb.create_sheet("Change Log")
    for i, h in enumerate(CHANGELOG_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(CHANGELOG_HEADERS), height=22)

    for r, row in enumerate(CHANGELOG, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = WRAP

    table = Table(displayName="ChangeLog", ref=f"A1:{get_column_letter(len(CHANGELOG_HEADERS))}{len(CHANGELOG)+1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    ws.freeze_panes = "A2"
    autosize(ws, {"A": 12, "B": 22, "C": 60, "D": 40, "E": 20})
    ws.sheet_view.showGridLines = False
    return ws


def main():
    wb = Workbook()
    build_story_overview(wb)
    build_variables(wb)
    build_architecture(wb)
    build_glossary(wb)
    build_bugs(wb)
    build_changelog(wb)
    wb.active = 0
    wb.save(OUT_PATH)
    print(
        f"Wrote {OUT_PATH}: {len(STORY_OVERVIEW)} stories, {len(VARIABLES)} variables, "
        f"{len(GLOSSARY)} glossary terms, {len(BUGS)} known issues."
    )


if __name__ == "__main__":
    main()
